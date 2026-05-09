from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

EXCEL_FILE = 'registros_ewaste.xlsx'

@app.route('/')
def index():
    # Busca automáticamente el archivo HTML
    if os.path.exists('index.html'):
        return send_file('index.html')
    elif os.path.exists('ecorecycle (1).html'):
        return send_file('ecorecycle (1).html')
    elif os.path.exists('ecorecycle.html'):
        return send_file('ecorecycle.html')
    else:
        return "❌ No se encuentra ningún archivo HTML en la carpeta"

@app.route('/guardar', methods=['POST'])
def guardar():
    try:
        datos = request.json
        print("📥 Recibido:", datos)
        
        # Crear Excel si no existe
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Fecha", "Tipo", "Marca", "Peso", "Cantidad", "Descripcion", "Punto", "Estado"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        
        nuevo_id = ws.max_row
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            nuevo_id,
            fecha,
            datos.get('tipo', ''),
            datos.get('marca', ''),
            datos.get('peso', 0),
            datos.get('cantidad', 1),
            datos.get('descripcion', ''),
            datos.get('punto', ''),
            'Pendiente'
        ])
        
        wb.save(EXCEL_FILE)
        print(f"✅ Guardado en Excel - ID: {nuevo_id}")
        
        return jsonify({"ok": True, "id": nuevo_id})
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"ok": False, "error": str(e)})

@app.route('/listar', methods=['GET'])
def listar():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify([])
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        registros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                registros.append({
                    'id': row[0],
                    'fecha': row[1] if row[1] else '',
                    'tipo': row[2] if row[2] else '',
                    'marca': row[3] if row[3] else '',
                    'peso': row[4] if row[4] else 0,
                    'cantidad': row[5] if row[5] else 1,
                    'descripcion': row[6] if row[6] else '',
                    'punto': row[7] if row[7] else '',
                    'estado': row[8] if row[8] else 'Pendiente'
                })
        
        return jsonify(registros)
        
    except Exception as e:
        print(f"❌ Error al listar: {e}")
        return jsonify([])

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 SERVIDOR ECORECYCLE")
    print(f"📁 Carpeta actual: {os.getcwd()}")
    print("📄 Archivos HTML encontrados:")
    for archivo in os.listdir('.'):
        if archivo.endswith('.html'):
            print(f"   - {archivo}")
    print(f"📁 Excel: {EXCEL_FILE}")
    print("🌐 Abre: http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=True, host='localhost', port=5000)