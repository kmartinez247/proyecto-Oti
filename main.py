from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)
nombre_archivo = 'registros_ewaste.xlsx'

# Configuración inicial del Excel
if not os.path.exists(nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.append(["Aparato", "Cantidad", "Categoría", "Estado", "Destino"])
    wb.save(nombre_archivo)

@app.route('/')
def index():
    # Carga tu archivo ecorecycle (1).html
    return render_template('ecorecycle (1).html')

@app.route('/enviar', methods=['POST'])
def enviar():
    # Captura los datos del formulario HTML
    aparato = request.form.get('aparato')
    cantidad = request.form.get('cantidad')
    categoria = request.form.get('categoria')
    estado = request.form.get('estado')

    # Tu lógica de clasificación de ingeniería
    if "Monitor" in aparato or "Pantalla" in aparato:
        destino = "Reciclaje Especializado"
    elif "Funciona" in estado:
        destino = "Laboratorio UNAB"
    else:
        destino = "Centro de Acopio General"

    # Guardar en Excel
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    ws.append([aparato, cantidad, categoria, estado, destino])
    wb.save(nombre_archivo)

    return "Registro exitoso. Datos guardados en Excel."

if __name__ == '__main__':
    app.run(debug=True)