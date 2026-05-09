"""
Módulo unificado para manejar la base de datos registros_ewaste.xlsx
Todos los componentes del proyecto usan este mismo archivo
"""

from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

DB_FILE = 'registros_ewaste.xlsx'

def get_workbook():
    """Obtiene el workbook de Excel"""
    if os.path.exists(DB_FILE):
        return load_workbook(DB_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros E-Waste"
        # Cabeceras unificadas
        headers = ["ID", "Fecha", "Usuario", "Aparato", "Marca", 
                   "Cantidad/Peso", "Categoria", "Estado", "Destino", 
                   "Punto Entrega", "Comentario"]
        ws.append(headers)
        wb.save(DB_FILE)
        return wb

def guardar_residuo(aparato, cantidad_peso, categoria, estado, 
                    comentario="", usuario="usuario", destino="", 
                    punto_entrega=""):
    """Guarda un nuevo registro en la base de datos"""
    wb = get_workbook()
    ws = wb.active
    
    # Obtener próximo ID
    next_id = ws.max_row
    
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Determinar destino según reglas
    if not destino:
        if "Monitor" in aparato or "Pantalla" in aparato or "TV" in aparato:
            destino = "Reciclaje Especializado"
        elif "Funciona" in estado or "Bueno" in estado:
            destino = "Laboratorio UNAB"
        else:
            destino = "Centro de Acopio General"
    
    nuevo_registro = [
        next_id, fecha, usuario, aparato, "", 
        cantidad_peso, categoria, estado, destino, 
        punto_entrega, comentario
    ]
    
    ws.append(nuevo_registro)
    wb.save(DB_FILE)
    return next_id

def obtener_todos_registros():
    """Obtiene todos los registros de la base de datos"""
    wb = get_workbook()
    ws = wb.active
    
    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Si tiene ID
            registros.append({
                'id': row[0],
                'fecha': row[1],
                'usuario': row[2],
                'aparato': row[3],
                'marca': row[4],
                'cantidad_peso': row[5],
                'categoria': row[6],
                'estado': row[7],
                'destino': row[8],
                'punto_entrega': row[9],
                'comentario': row[10]
            })
    return registros

def obtener_estadisticas():
    """Obtiene estadísticas para el dashboard"""
    registros = obtener_todos_registros()
    
    total_peso = 0
    total_dispositivos = 0
    categorias = {}
    destinos = {}
    
    for r in registros:
        # Intentar convertir peso a float
        try:
            peso = float(r['cantidad_peso']) if r['cantidad_peso'] else 0
        except (ValueError, TypeError):
            peso = 0
        
        total_peso += peso
        total_dispositivos += 1
        
        cat = r['categoria'] or 'Sin categoría'
        categorias[cat] = categorias.get(cat, 0) + peso
        
        dest = r['destino'] or 'Sin destino'
        destinos[dest] = destinos.get(dest, 0) + 1
    
    return {
        'total_peso': total_peso,
        'total_dispositivos': total_dispositivos,
        'categorias': categorias,
        'destinos': destinos
    }

def eliminar_registro(id_registro):
    """Elimina un registro por ID"""
    wb = get_workbook()
    ws = wb.active
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == id_registro:
            ws.delete_rows(row_idx)
            break
    
    wb.save(DB_FILE)

def actualizar_estado(id_registro, nuevo_estado):
    """Actualiza el estado de un registro"""
    wb = get_workbook()
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):
        if row[0].value == id_registro:
            row[7].value = nuevo_estado  # Columna Estado
            break
    
    wb.save(DB_FILE)